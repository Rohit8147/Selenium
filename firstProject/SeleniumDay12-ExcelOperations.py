import time
import openpyxl


from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

service_obj= Service("C://Users//a778983//OneDrive - Eviden//Desktop//chromedriver-win64//chromedriver.exe")
driver=webdriver.Chrome(service=service_obj)
driver.maximize_window()
driver.get("https://www.durgasoft.com/registration.asp")
driver.implicitly_wait(5)

workbook=openpyxl.load_workbook("C://Users//a778983//OneDrive - Eviden//Desktop//SampleInfo.xlsx")
sheet=workbook.get_sheet_by_name("Sheet1")
print(sheet.cell(3,1).value)

Name=sheet.cell(3,1).value
Phone=sheet.cell(3,2).value
driver.find_element(By.ID,"name").send_keys(Name)
driver.find_element(By.ID,"ph_no").send_keys(Phone)

time.sleep(5)

row=sheet.max_row
column=sheet.max_column
print(row)
print(column)
for r in range(1,row+1):
    for c in range(1,column+1):
        print((sheet.cell(r,c).value), end = " ")
    print()