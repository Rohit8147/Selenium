import unittest
import openpyxl
from symtable import Class

from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import time



class UnitTestScripts(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        service_obj = Service("C://Users//a778983//OneDrive - Eviden//Desktop//chromedriver-win64//chromedriver.exe")
        global driver
        driver = webdriver.Chrome(service=service_obj)
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        driver.maximize_window()
        print("Start")
        driver.implicitly_wait(5)

    def test1_login(self):
        workbook = openpyxl.load_workbook("C://Users//a778983//OneDrive - Eviden//Desktop//SampleInfo.xlsx")
        sheet = workbook.get_sheet_by_name("Sheet1")
        Name = sheet.cell(2, 1).value
        Pass = sheet.cell(2, 2).value
        driver.find_element(By.XPATH,"//input[@name='username']").send_keys(Name)
        driver.find_element(By.XPATH, "//input[@name='password']").send_keys(Pass)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        print("login Done")
        time.sleep(4)

    def test2_Employee(self):
        driver.find_element(By.XPATH, "//span[text()='PIM']").click()
        driver.find_element(By.XPATH, "//a[text()='Add Employee']").click()
        workbook = openpyxl.load_workbook("C://Users//a778983//OneDrive - Eviden//Desktop//SampleInfo.xlsx")
        sheet = workbook.get_sheet_by_name("Sheet1")
        First = sheet.cell(7, 1).value
        Middle = sheet.cell(7, 2).value
        Last = sheet.cell(7, 3).value
        driver.find_element(By.XPATH,"//input[@name='firstName']").send_keys(First)
        driver.find_element(By.XPATH, "//input[@name='middleName']").send_keys(Middle)
        driver.find_element(By.XPATH, "//input[@name='lastName']").send_keys(Last)
        driver.find_element(By.XPATH, "//button[text()=' Save ']").click()

        time.sleep(10)




    def test3_Search(self):

        driver.find_element(By.XPATH, "//a[text()='Employee List']").click()
        print("test3")
        workbook = openpyxl.load_workbook("C://Users//a778983//OneDrive - Eviden//Desktop//SampleInfo.xlsx")
        sheet = workbook.get_sheet_by_name("Sheet1")
        First = sheet.cell(7, 1).value
        Middle = sheet.cell(7, 2).value
        time.sleep(5)

        driver.find_element(By.XPATH, "(//input[@placeholder='Type for hints...'])[1]").send_keys(First)
        print("enter name")
        driver.find_element(By.XPATH, "//button[text()=' Search ']").click()
        time.sleep(4)
        sum1 = First + " " + Middle
        print(First + " " + Middle)

        l1 = driver.find_elements(By.XPATH, "//div[@class='oxd-table-cell oxd-padding-cell']")
        if l1[1].text == sum1:
            print("It's correct")

        else:
            print("It's wrong")




    @classmethod
    def tearDownClass(cls):
        driver.close()
        print("Close")

